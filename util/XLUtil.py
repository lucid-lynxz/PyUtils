# !/usr/bin/env python
# -*- coding:utf-8 -*-
import copy
import os
import re
import time

import openpyxl
import openpyxl.cell
import openpyxl.utils
import openpyxl.workbook
from openpyxl.worksheet import cell_range

from util.IExcelUtil import IExcelUtil
from util.FileUtil import FileUtil
from util.CommonUtil import CommonUtil


class XLUtil(IExcelUtil):
    """
    openpyxl 工具类
    需要安装包: pip install openpyxl
    """

    def __init__(self, xlPath):
        """
        :param xlPath: excel文件绝对路径
        """
        self.filePath = os.path.abspath(xlPath.strip())
        self.workBook = None
        self.workBookDataOnly = None
        # XLUtil.justOpenByExcel(self.filePath)
        self.reload()
        CommonUtil.printLog("xlutil init finished %s" % self.filePath)

    def reload(self):
        self.save()
        self.workBookDataOnly = openpyxl.load_workbook(self.filePath, data_only=True)
        self.workBook = openpyxl.load_workbook(self.filePath)

    def getAllSheetNames(self) -> list:
        """
        获取所有表名, 是个list
        :return:
        """
        return self.workBook.sheetnames

    def addSheet(self, title: str = None, afterActive: bool = True, autoSave: bool = True) -> openpyxl.worksheet:
        """
        在当前表格后添加一个新表格
        :param title: 表格名, 若不传, 则自动按照excel规则生成 sheet, sheet1, sheet2....
        :param afterActive: True-在当前表格后面新建 False-在当前表格前面新建
        :param autoSave: 创建后是否自动保存,默认True
        :return:新建的表格对象
        """
        activeName = self.workBook.active.title  # 当前sheet名
        activeIndex = self.getAllSheetNames().index(activeName)
        # CommonUtil.printLog("activeName=%s,index=%s,title=%s" % (activeName, activeIndex, title))
        newSheet = self.workBook.create_sheet(title, activeIndex + 1 if afterActive else max(0, activeIndex))
        CommonUtil.printLog(
            "addSheet title=%s, afterActive=%s,activeName=%s" % (newSheet.title, afterActive, activeName))
        if autoSave:
            self.save()
        return newSheet

    @staticmethod
    def colNum2Str(colIndex: int) -> str:
        """
        行列数字转字母, 如 1 -> "A", 27 -> "AA"
        :param colIndex: 列数字
        :return: 对应的列号字母
        """
        return openpyxl.utils.get_column_letter(colIndex)

    @staticmethod
    def getRangeWH(rangeStr: str) -> tuple:
        """
        计算单元格区域的宽高
        :param rangeStr: 单元格地址,如: "A1:Z2" 若不包含冒号,则直接返回 (1,1)
        :return: tuple(colCnt,rowCnt) colCnt表示列数, rowCnt表示行数
        """
        if ":" not in rangeStr:
            return 1, 1
        ltColIndex, ltRowIndex = XLUtil.getColRowIndex(rangeStr)
        brCellAddress = rangeStr.split(":")[1]
        brColIndex, brRowIndex = XLUtil.getColRowIndex(brCellAddress)
        return brColIndex - ltColIndex + 1, brRowIndex - ltRowIndex + 1

    @staticmethod
    def getColNum(colRowStr: str) -> int:
        """
        列字母转数字, 如: "A" -> 1, "AA" -> 27, "aa" -> 27
        :param colRowStr: 字母信息,如: "AA" 或者 "A12" 对于后者,会自动提取其中的字母部分,并将其转换为数字
        :return:
        """
        colStr = re.findall(r'[a-zA-Z]+', colRowStr)[0]  # 提取列号字母,行号都是数字
        return openpyxl.utils.column_index_from_string(colStr)

    @staticmethod
    def getRowNum(colRowStr: str) -> int:
        """
        获取行号数字
        :param colRowStr:行列信息,如: "A1" ,则提取数字1
        :return:
        """
        return int(colRowStr.replace(re.findall(r'[a-zA-Z]+', colRowStr)[0], ""))

    @staticmethod
    def colRowNum2Str(colNum: int, rowNum: int) -> str:
        """
        将行/列数字转换为字符串
        如: (1,1) -> "A1"
        :param colNum: 列号,从1开始,1表示 "A" 列
        :param rowNum: 行号,从1开始
        :return:
        """
        return "%s%s" % (XLUtil.colNum2Str(colNum), rowNum)

    @staticmethod
    def getColRowIndex(colRowStr: str) -> tuple:
        """
        将单元格字符串地址转为行列号tuple
        :param colRowStr: 单元格地址,如: "A1", "A1:Z1" 返回左上角单元格序号
        :return: tuple (colNum, rowNum), 如:"A1" -> 1,1
        """
        colRowStr = colRowStr.split(":")[0].split(",")[0]
        return XLUtil.getColNum(colRowStr), XLUtil.getRowNum(colRowStr)

    @staticmethod
    def shiftColRowIndex(colRowOrRangeStr: str, deltaCol: int, deltaRow: int) -> str:
        """
        对单元格地址(如: A1)或者单元格区域地址(如: A1:Z1) 进行偏移后得到最终的地址字符串
        示例: A1, 列偏移1, 行偏移1 --> B2
             A1:Z1 --> B2:AA2
        :param colRowOrRangeStr: 原单元格或者单元格区域地址,如: "A1" 或 "A1:Z1"
        :param deltaCol: 列偏移量, +1 表示往右偏移, -1 表示往左偏移
        :param deltaRow: 行偏移量, +1 表示往下偏移, -1 表示往上偏移
        :return: 偏移后的地址
        """
        ltCol, ltRow = XLUtil.getColRowIndex(colRowOrRangeStr)
        result = "%s%d" % (XLUtil.colNum2Str(ltCol + deltaCol), ltRow + deltaRow)
        if ":" in colRowOrRangeStr:
            brCol, brRow = XLUtil.getColRowIndex(colRowOrRangeStr.split(":")[1])
            result = "%s:%s%s" % (result, XLUtil.colNum2Str(brCol + deltaCol), brRow + deltaRow)
        return result

    def getMergeCellsInRange(self, srcSheetTitle: str, colRowRangeStr: str) -> set:
        """
        提取指定范围内的所有合并单元格信息
        :param srcSheetTitle: 表格名称
        :param colRowRangeStr: 待搜索的范围
        :return: set() 元素是合并区域的坐标字符串, 如: "A1:X1", "B2:B10"
        """
        srcSheet = self.workBook[srcSheetTitle]
        allMergeCells = srcSheet.merged_cells  # 表格中所有的合并单元格

        # 处理合并单元格
        srcSheet = self.workBook[srcSheetTitle]
        srcRange = srcSheet[colRowRangeStr]
        mergeCellsInSrcRange = set()
        for tRow in srcRange:
            for tCell in tRow:
                if isinstance(tCell, openpyxl.cell.MergedCell):
                    for mCellRange in allMergeCells:
                        if mCellRange.__contains__(tCell.coordinate):
                            mergeCellsInSrcRange.add("%s%s:%s%s" % (
                                XLUtil.colNum2Str(mCellRange.min_col), mCellRange.min_row,
                                XLUtil.colNum2Str(mCellRange.max_col), mCellRange.max_row))
        # CommonUtil.printLog(mergeCellsInSrcRange)
        return mergeCellsInSrcRange

    @staticmethod
    def copyCellFormat(srcCell: openpyxl.cell, dstCell: openpyxl.cell):
        """
        复制单元格格式到 dstCell 中
        :param srcCell: 源单元格
        :param dstCell: 目标单元格
        :return:
        """
        # 过滤掉合并单元格,否则报错: AttributeError: 'MergedCell' object attribute 'data_type' is read-only
        if not isinstance(dstCell, openpyxl.cell.MergedCell):
            dstCell.data_type = srcCell.data_type
            # if srcCell.comment:
            dstCell.comment = copy.copy(srcCell.comment)

        dstCell.fill = copy.copy(srcCell.fill)
        if srcCell.has_style:
            dstCell._style = copy.copy(srcCell._style)
            dstCell.font = copy.copy(srcCell.font)
            dstCell.border = copy.copy(srcCell.border)
            dstCell.fill = copy.copy(srcCell.fill)
            dstCell.number_format = copy.copy(srcCell.number_format)
            dstCell.protection = copy.copy(srcCell.protection)
            dstCell.alignment = copy.copy(srcCell.alignment)

        if srcCell.hyperlink:
            dstCell._hyperlink = copy.copy(srcCell.hyperlink)

    def getCellValue(self, sheetTitle: str, cellAddress: str, dataOnly: bool = True) -> str:
        """
        获取指定单元格的内容
        :param sheetTitle: 工作表名称
        :param cellAddress: 单元格地址,如: "A1" 或者 "A1:B2" 取左上角单元格
        :param dataOnly: True-若是公式,则提取计算结果
        :return: 单元格内容字符串
        """
        colIndex, rowIndex = XLUtil.getColRowIndex(cellAddress)
        sheet = self.workBookDataOnly[sheetTitle] if dataOnly else self.workBook[sheetTitle]
        return sheet.cell(row=rowIndex, column=colIndex).value

    def copyValue(self, srcSheetTitle: str,
                  srcColRowRangeStr: str,
                  dstSheetTitle: str,
                  dstColRowStrStart: str = None,
                  withFormat: bool = True,
                  dataOnly: bool = False,
                  autoSave: bool = True):
        """
        复制源表格区域内容到目标表格中
        :param srcSheetTitle:  源表格名称
        :param srcColRowRangeStr: 待复制的区域, 如: "A1:D3"
        :param dstSheetTitle: 目标表格名称
        :param dstColRowStrStart: 目标表格粘贴区域的左上角单元格,如: "A1",默认表示与 srcColRowRangeStr 相同
        :param withFormat: 是否需要复制单元格格式,默认True
        :param dataOnly: 是否只复制单元格值 True-复制结果值 False-若有公式,则复制的是公式
        :param autoSave: 复制完成后是否自动保存,默认 True
        :return:
        """

        # 若目标表格不存在,则新建一个
        if dstSheetTitle is None or dstSheetTitle not in self.getAllSheetNames():
            dstSheetTitle = self.addSheet(dstSheetTitle).title

        if dstColRowStrStart is None or len(dstColRowStrStart) == 0:
            dstColRowStrStart = srcColRowRangeStr.split(":")[0]
            pass

        mergeCellSet = self.getMergeCellsInRange(srcSheetTitle, srcColRowRangeStr)
        srcSheet = self.workBook[srcSheetTitle]  # 原始格式
        dataOnlySrcSheet = self.workBookDataOnly[srcSheetTitle]  # 仅包含结果值
        dstSheet = self.workBook[dstSheetTitle]

        # 根据 dstColRowStrStart 进行偏移
        srcColIndex, srcRowIndex = XLUtil.getColRowIndex(srcColRowRangeStr)
        dstColIndex, dstRowIndex = XLUtil.getColRowIndex(dstColRowStrStart)
        deltaCol = dstColIndex - srcColIndex
        deltaRow = dstRowIndex - srcRowIndex

        # 处理合并单元格,并复制其内容
        for cellRangeStr in mergeCellSet:
            dstSheet.merge_cells(XLUtil.shiftColRowIndex(cellRangeStr, deltaCol, deltaRow))  # 合并单元格区域
            # srcCr = cell_range.CellRange(range_string=cellRangeStr)
            # srcValue = srcSheet.cell(srcCr.min_row, srcCr.min_col).value
            # dstSheet.cell(srcCr.min_row, srcCr.min_col).value = srcValue

        # 复制单元格内容和宽高
        srcCr = cell_range.CellRange(range_string=srcColRowRangeStr)
        for colIndex in range(srcCr.min_col, srcCr.max_col + 1):
            # 复制列宽
            srcColWidth = srcSheet.column_dimensions[XLUtil.colNum2Str(colIndex)].width
            dstSheet.column_dimensions[XLUtil.colNum2Str(colIndex + deltaCol)].width = srcColWidth
            for rowIndex in range(srcCr.min_row, srcCr.max_row + 1):
                # 复制行高
                if colIndex == srcCr.min_col:
                    srcRowHeight = srcSheet.row_dimensions[rowIndex].height
                    dstSheet.row_dimensions[rowIndex + deltaRow].height = srcRowHeight

                srcCell = srcSheet.cell(rowIndex, colIndex)
                dstCell = dstSheet.cell(rowIndex + deltaRow, colIndex + deltaCol)
                dataOnlySrcCell = dataOnlySrcSheet.cell(rowIndex, colIndex)

                # 复制单元格格式
                if withFormat:
                    XLUtil.copyCellFormat(srcCell, dstCell)

                # 复制单元格内容
                # 若是合并单元格,则跳过(左上角单元格判断为false,会复制,其余逃过)
                if not isinstance(dstCell, openpyxl.cell.MergedCell):
                    dstCell.value = dataOnlySrcCell.value if dataOnly else srcCell.value
        if autoSave:  # 保存后才生效
            self.save()

    def save(self, filePath: str = None):
        """
        保存修改到文件中
        :param filePath:保存路径,若传入None,则表示保存,否则是另存
        :return:
        """
        try:
            if self.workBook is not None:
                self.workBook.save(self.filePath if filePath is None else filePath)
        except Exception as e:
            CommonUtil.printLog(f"XLUtil.save error: {e}")

    def close(self, save: bool = True, filePath: str = None):
        """
        关闭excel
        :param save: 是否保存
        :param filePath: 保存路径，默认为原路径
        :return:
        """
        if save:
            self.save(filePath)

        if self.workBook is not None:
            self.workBook.close()
            self.workBook = None

        if self.workBookDataOnly is not None:
            self.workBookDataOnly.close()
            self.workBookDataOnly = None

    @staticmethod
    def justOpenByExcel(xlFilePath: str):
        """
        仅通过pywin32库用excel程序打开并关闭excel文件
        参考: https://blog.csdn.net/claria029/article/details/116486904
        :param xlFilePath:
        :return:
        """
        # 仅系统是win时才可用
        import platform
        if "Windows" != platform.system():
            return
        from win32com.client import Dispatch
        xlApp = Dispatch("Excel.Application")
        xlApp.Visible = False
        xlApp.DisplayAlerts = False  # 是否显示警告
        xlBook = xlApp.Workbooks.Open(xlFilePath)
        xlBook.Save()
        xlBook.Close()

    def getMaxRowNumOfColumn(self, colLetter: str, sheetName: str = None) -> int:
        """
        根据所给列号(数字或字母),返回该列有数据的最大行号
        注意: 列最后的一个有数据的单元格若是合并单元格,则只会返回第一个单元格行号
        :param colLetter: 目标列号,字母形式, 如: 'A'
        :param sheetName: 表格名,若为空,则表示当前活动工作表
        """
        if sheetName is None or len(sheetName) == 0:
            srcSheet = self.workBook.active  # 当前活动工作表
        else:
            srcSheet = self.workBook[sheetName]

        colNum = self.getColNum(colLetter)
        maxRow: int = 0
        for r in range(srcSheet.max_row, srcSheet.min_row, -1):
            c = srcSheet.cell(r, colNum)
            if c.value is not None:
                maxRow = max(maxRow, c.row)
                break
        return maxRow

    @staticmethod
    def take_shot(xlPath: str, sheetName: str, area: str, imgPath: str = None, continueMode: bool = False) -> tuple:
        """
        python excel截图， 参考： https://blog.csdn.net/weixin_30378623/article/details/96566903
        需要安装win库：
        pip install pywin32 -i https://pypi.douban.com/simple
        pip install pillow -i https://pypi.douban.com/simple
        :param xlPath: excel源文件路径
        :param sheetName: 待截图的表格名
        :param area:待截图的区域，如: "A1:X2"
        :param imgPath: 图片保存的绝对路径
        :param continueMode: 若为True,则表示若图片已经存在,则不重新截图
        :return: (bool,str) 依次表示是否截图成功，以及图片的位置
        """
        # 仅系统是win时才可用
        import platform
        if "Windows" != platform.system():
            return False, imgPath

        from util.ExcelScreenShotUtil import ExcelTakeShotUtil
        tXlPath = os.path.abspath(xlPath)
        imgPath = FileUtil.recookPath(imgPath)

        _shot_util = ExcelTakeShotUtil(tXlPath, sheetName)
        result = _shot_util.take_shot(area, imgPath, continueMode)
        _shot_util.close()
        return result


if __name__ == "__main__":
    xlPath = "/Users/lynxz/Desktop/Book1.xlsx"
    xl_util = XLUtil(xlPath)
    sheetTitle = "测试"
    CommonUtil.printLog(xl_util.workBook[sheetTitle]["F5"].value)
    CommonUtil.printLog(xl_util.workBookDataOnly[sheetTitle]["F5"].value)

    CommonUtil.printLog(xl_util.getAllSheetNames())
    sheet = xl_util.workBook.active  # 当前活动工作表
    CommonUtil.printLog("activeSheet=%s,rows=%s, columns=%s,maxRowOfColumnA=%s" % (sheet.title, sheet.rows,
                                                                                   sheet.columns,
                                                                                   xl_util.getMaxRowNumOfColumn(
                                                                                       'A')))  # 行列数据生成器
    CommonUtil.printLog("maxRow=%s,maxColumns=%s" % (sheet.max_row, sheet.max_column))  # 表行数/列数 (数据区域)
    CommonUtil.printLog("sheet merge cells %s" % sheet.merged_cells.ranges)

    # 迭代获取行数据
    for row in sheet.rows:
        line = [cell.value for cell in row]
        CommonUtil.printLog(line)
        break

    CommonUtil.printLog("A1.value=%s" % sheet["A1"].value)

    # column字母转数字
    CommonUtil.printLog(
        "colStr2Num A=%s, a=%s, A1=%s, a1=%s, AA=%s, aa=%s" % (xl_util.getColNum("A"), xl_util.getColNum("a"),
                                                               xl_util.getColNum("A1"), xl_util.getColNum("a1"),
                                                               xl_util.getColNum("AA"), xl_util.getColNum("aa")))
    # column数字转字母
    CommonUtil.printLog("colNum2Str 1=%s, 27=%s" % (xl_util.colNum2Str(1), xl_util.colNum2Str(27)))

    # 提取行号
    CommonUtil.printLog("getRowNum A1=%s, AA138=%s" % (xl_util.getRowNum("A1"), xl_util.getRowNum("AA138")))

    # 合并单元格判断
    sheetTitle = "综治、禁毒、卫生协管、安全协管、党建协管（不含陈舜乐）"
    dstSheetTitle = "Sheet"
    # xl_util.copyValue(sheetTitle, "A1:Z7", "sheet1", None, withFormat=True)
    # xl_util.copyValue(sheetTitle, "A1:Z7", "sheet2", None, withFormat=False)
    # xl_util.copyValue(sheetTitle, "A1:Z7", "sheet3", "B2", withFormat=True)
    # xl_util.copyValue(sheetTitle, "A1:A1", "sheet4", "B2", withFormat=True)
    # xl_util.workBook[dstSheetTitle].merge_cells(range_string="B1:B10")
    # xl_util.save()

    # CommonUtil.printLog("getColRowIndex(A1)=%d,%d" % XLUtil.getColRowIndex("A1"))
    # CommonUtil.printLog("getColRowIndex(A1:z3)=%d,%d" % XLUtil.getColRowIndex("A1:Z3"))
    # CommonUtil.printLog("getColRowIndex(A1,A2)=%d,%d" % XLUtil.getColRowIndex("A1,A2"))
    # CommonUtil.printLog("getColRowIndex(A1:Z3,A2:B5)=%d,%d" % XLUtil.getColRowIndex("A1:Z3,A2:B5"))
    # CommonUtil.printLog("shiftColRowIndex(A1)=%s" % XLUtil.shiftColRowIndex("A1", 1, 1))
    # CommonUtil.printLog("shiftColRowIndex(A1:Z1)=%s" % XLUtil.shiftColRowIndex("A1:Z1", 1, 1))
    #
    # xl_util.addSheet(afterActive=True) # 在当前表格后新增一个表格
    # xl_util.addSheet(afterActive=False) # 在当前表格前插入一个表格
    # CommonUtil.printLog(xl_util.getAllSheetNames()) # 获取所有表格名

    CommonUtil.printLog("A3=%s" % xl_util.getCellValue(sheetTitle, "A3"))
    CommonUtil.printLog("A5=%s" % xl_util.getCellValue(sheetTitle, "A5"))
    CommonUtil.printLog("Y5=%s" % xl_util.getCellValue(sheetTitle, "Y5"))
