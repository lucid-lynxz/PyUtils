# !/usr/bin/env python
# -*- coding:utf-8 -*-
"""
工资单汇总表拆分,每个人的工资信息加上固定标题行
"""
import os
import sys
import time

from openpyxl.worksheet import cell_range, pagebreak

# 将整个根目录加入到path环境变量中,并且是第一个,以便脚本导包时柯使用绝对路径
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from util.FileUtil import FileUtil
from util.ConfigUtil import NewConfigParser
from util.TimeUtil import TimeUtil, log_time_consume
from util.CommonUtil import CommonUtil
from util.XLUtil import XLUtil
from util.MailUtil import MailUtil
from util.NetUtil import NetUtil
from base.BaseConfig import BaseConfig
from util.ExcelScreenShotUtil import ExcelTakeShotUtil


class EmployeeSalaryInfo(object):
    """
    员工工资截图信息对象
    """

    def __init__(self, imagePath: str, name: str = None, email: str = None):
        """
        员工信息
        :param imagePath: 截图文件路径, 文件名称格式为: 姓名_邮箱地址.png
        :param name: 员工姓名, 若为空, 则通过截图文件名截取,如: 张三
        :param email: 员工邮箱地址, 若为空,则通过截图文件名截取, 如: 123456@qq.com
        """
        self.imagePath: str = imagePath  # 工资截图地址
        fullName, fileNameOly, ext = FileUtil.getFileName(imagePath)
        self.imageName: str = fullName  # 图片名称,带后缀, 如: 张三_123456@qq.com.png

        arr = fileNameOly.split('_')
        _name: str = arr[0]
        _email: str = '_'.join(arr[1:])
        self.name: str = _name if CommonUtil.isNoneOrBlank(name) else name
        self.email: str = _email if CommonUtil.isNoneOrBlank(email) else email


class SendSalaryEmailImpl(BaseConfig):
    """
    自动拆分工资条,截图,发送邮件实现类
    """
    xl_path: str = ''  # excel文件路径
    imgDirPath: str = ''  # 截图缓存目录
    success_pic_cnt: int = 0  # 截图成功的数量, 于 employeeCnt 相等时才允许发送邮件
    fail_pic_cnt: int = 0  # 截图失败的数量, 等于0时才允许发送邮件

    def onRun(self):
        # 1. 确定工资单总表excel文件路径
        configSection = self.configParser.getSectionItems('config')
        NetUtil.robot_dict = self.configParser.getSectionItems('robot')  # 推送消息设置

        salaryPath: str = FileUtil.recookPath(configSection['salaryPath'])
        CommonUtil.printLog(f"当前config.ini中指定的salaryPath={FileUtil.absPath(salaryPath)}")

        while not FileUtil.isFileExist(salaryPath):
            salaryPath = CommonUtil.get_input_info("salaryPath不存在,请输入工资表excel文件路径:", "")

        sheetName: str = configSection['sheetName']
        needConfirm: bool = configSection['needConfirm'].lower() == 'yes'  # 选择sheet名或者发送邮件前是否需要二次确认,yes/no, 默认: yes
        xlPath: str = salaryPath  # 目标excel文件路径

        if FileUtil.isDirFile(salaryPath):
            # 1.1 提取目录下的所有excel文件
            subExcelList: list = FileUtil.get_sub_file_names(salaryPath, ['.xlsx'])
            excelCnt = len(subExcelList)

            # 1.2 若目录下没有excel文件,则提示用户输入
            # 若目录下有多个excel文件,则提示用户选择其中一个
            needInput: bool = False
            if excelCnt > 1:
                CommonUtil.printLog("存在多个excel文件: ")
                for index, name in enumerate(subExcelList):
                    CommonUtil.printLog("\t%d\t%s" % (index, name))
                excelIndex = int(CommonUtil.get_input_info("请选择序号(默认:0): ", "0"))
                xlPath = f"{salaryPath}/{subExcelList[excelIndex]}"
            elif excelCnt == 1:
                CommonUtil.printLog(f"指定目录下只有一个excel文件: {subExcelList[0]}")
                if not needConfirm or "yes" == CommonUtil.get_input_info("是否使用该文件 yes/no (默认:yes): ", "yes"):
                    xlPath = f"{salaryPath}/{subExcelList[0]}"
                else:
                    needInput = True
            else:
                needInput = True

            if needInput:
                xlPath = input("请输入excel文件路径: ").strip()

            # 判断文件是否存在
            if not FileUtil.isFileExist(xlPath):
                CommonUtil.printLog(f"excel文件不存在, 请检查 {xlPath}")
                exit(1)

        # 2. 确定工资单sheet名称
        self.xl_path = xlPath
        _xlutil = XLUtil(xlPath)
        allSheetNames = _xlutil.getAllSheetNames()

        needInput = False
        sheetTitle = sheetName
        CommonUtil.printLog("其sheet名列表为: ", includeTime=False)
        nameIndexList = list(enumerate(allSheetNames))
        for index, name in nameIndexList:
            CommonUtil.printLog("\t%d\t%s" % (index, name), includeTime=False)

        if sheetName in allSheetNames:
            needInput = needConfirm and "no" == CommonUtil.get_input_info(
                f"是否使用'{sheetName}', yes/no (默认:yes):  ", "yes")
        else:
            needInput = True

        if needInput:
            sheetTitleIndex = int(CommonUtil.get_input_info("请输入要切分的工作表名序号(默认:0): ", "0"))
            sheetTitle = nameIndexList[sheetTitleIndex][1]
        CommonUtil.printLog("工资汇总表名为: %s" % sheetTitle)

        xlDirPath = os.path.dirname(self.xl_path)
        _, xlFileName, _ = FileUtil.getFileName(self.xl_path)
        self.imgDirPath = FileUtil.recookPath(f'{xlDirPath}/imgCache/{xlFileName}_{sheetTitle}/')  # 截图缓存目录

        # todo 提示用户是否要查看/修改配置(支持超时自动跳过)

        # 3. 进行工资条拆分并自动截图(仅windows支持)
        headerRangeStr: str = configSection['headerRangeStr']  # 标题区域范围,默认: A3:X4
        firstColIndex, firstRowIndex = XLUtil.getColRowIndex(headerRangeStr)  # 第一列序号, 第一行序号
        colCnt, rowCnt = XLUtil.getRangeWH(headerRangeStr)  # 标题栏宽高

        firstColLetter = _xlutil.colNum2Str(firstColIndex)  # 首列字母
        maxRowNum = _xlutil.getMaxRowNumOfColumn(firstColLetter, sheetTitle)  # 首列有数据的最大行号
        CommonUtil.printLog('  计算得到首列的最大行号为: %s' % maxRowNum)

        sumRowCnt: int = int(configSection['sumRowCnt'])  # 最后的'小计'行数,默认:1, 这几行不计入个人工资内容区域
        contentRangeSug = "%s%s:%s%s" % (  # 内容范围区域
            XLUtil.colNum2Str(firstColIndex), firstRowIndex + rowCnt,
            XLUtil.colNum2Str(firstColIndex + colCnt - 1), maxRowNum - sumRowCnt)
        totalSalaryContentRangeStr: str = contentRangeSug  # 工资内容区域范围
        CommonUtil.printLog('  计算得到总表工资内容区域为: %s' % totalSalaryContentRangeStr)

        eachSalaryContentRowCnt: int = int(configSection['eachSalaryContentRowCnt'])  # 个人工资内容行数,默认: 1
        eachGap: int = int(configSection['eachGap'])  # 工人工资信息之间的间隔行数,默认: 0
        addRowPageBreak: bool = configSection['addRowPageBreak'].lower() == 'yes'  # 是否为每人的工资条增加依次分页符,yes/no,默认: yes
        autoTakeShot: bool = configSection['autoTakeShot'].lower() == 'yes'  # 是否自动截图工资条信息,yes/no, 默认: yes
        autoSendMail: bool = configSection['autoSendMail'].lower() == 'yes'  # 截图后是否自动发送工资条信息到对方邮箱中,yes/no, 默认: yes
        emailColLetter: str = configSection['emailColLetter']  # 邮箱信息所在列号, 默认: Y
        nameColLetter: str = configSection['nameColLetter']  # 姓名所在列号, 默认: B
        employeeCnt: int = int(configSection['employeeCnt'])  # 员工数, 大于0有效, 发送邮件前需要图片数符合预期才新
        CommonUtil.printLog("参数配置完成,即将开始拆分,请稍后...\n\n")

        try:
            dstSheetTitle, dstRangeStr = self.split(_xlutil, sheetTitle, headerRangeStr, totalSalaryContentRangeStr,
                                                    None,
                                                    eachSalaryContentRowCnt, eachGap, addRowPageBreak)
            _xlutil.close()
            _xlutil = None
            NetUtil.push_to_robot(f'工资条拆分完成,请查看结果工作表:{dstSheetTitle}\nexcel文件:\n{self.xl_path}')

            if autoTakeShot:
                self.take_shot(dstSheetTitle, headerRangeStr, dstRangeStr, 1, emailColLetter, nameColLetter)

                # 4. 截图后进行邮件发送
                # CommonUtil.printLog("完成截图%s张, %s" % (len(imgMap), imgMap))
                if autoSendMail:
                    if employeeCnt <= 0 or employeeCnt == self.success_pic_cnt:
                        NetUtil.push_to_robot(f'工资条截图完成, 准备发送邮件, 共{self.success_pic_cnt}封')
                        self.sendEmailByImage(self.imgDirPath, self.configPath, needConfirm=needConfirm)
                    else:
                        msg = f'截图完成,失败:{self.fail_pic_cnt},成功:{self.success_pic_cnt},员工数:{employeeCnt}, 取消邮件发送,请重新截取'
                        NetUtil.push_to_robot(msg)
        except Exception as e:
            NetUtil.push_to_robot(f'拆分工资条发生异常, 请检查后重试: {e}')
        finally:
            if _xlutil is not None:
                _xlutil.close()

    @log_time_consume()
    def split(self, util: XLUtil, sheetTile: str, headerRangeStr: str, totalSalaryContentRangeStr: str,
              dstSheetTitle: str = None, eachSalaryContentRowCnt: int = 1, eachGap: int = 0,
              addRowPageBreak: bool = True) -> tuple:
        """
        拆分工资汇总表为每个人独立一个工资条(固定的标题栏+个人工资信息+gap空白分隔行)
        :param util: excel工具类, 如: XLUTIL("D:/salary.xlsx")
        :param sheetTile: 工资表名, 如: "sheet1"
        :param headerRangeStr:工资条固定标题栏区域,如: "A1:Z2"
        :param totalSalaryContentRangeStr: 员工工资表信息内容区域范围(剔除标题栏区域), 如 "A3:Z50"
        :param dstSheetTitle: 要复制到的目标表名, 默认为None,表示excel新建表格时自动命名
        :param eachSalaryContentRowCnt: 每个人的工资信息行数,默认1行
        :param eachGap: 个人工资标题栏和工资信息复制后, 空白行数,用于间隔,默认0
        :param addRowPageBreak:每复制一个工资条,增加一个打印分页符
        :return: (str,str):  最终的工作表名 以及 整体范围
        """
        dstSheet = util.addSheet(dstSheetTitle, False)
        dstSheetTitle = dstSheet.title
        headWidth, headHeight = XLUtil.getRangeWH(headerRangeStr)

        srcCr = cell_range.CellRange(range_string=totalSalaryContentRangeStr)
        totalRow = srcCr.max_row - srcCr.min_row + 1  # 总行数
        infoCnt = int(totalRow / eachSalaryContentRowCnt)  # 总信息条数
        copyInfoCnt = 0  # 已复制的信息条数
        copyRowCnt = 0  # 已复制的总行数
        takeScreenShotFailCnt = 0  # 截图失败的数量
        for rowIndex in range(srcCr.min_row, srcCr.max_row + 1, eachSalaryContentRowCnt):
            copyInfoCnt += 1
            # 添加分页符
            if addRowPageBreak:
                dstSheet.row_breaks.append(pagebreak.Break(id=copyRowCnt))

            # 复制标题栏
            CommonUtil.printLog("%s 正在复制第 %s/%s 条信息" % (TimeUtil.getTimeStr(), copyInfoCnt, infoCnt))
            startRowIndex = copyRowCnt + 1
            startCell = "A%s" % startRowIndex  # 起始单元格, 都从A列开始

            util.copyValue(sheetTile, headerRangeStr, dstSheetTitle, startCell, dataOnly=True, autoSave=False)
            copyRowCnt += headHeight

            # 复制个人工资条信息
            ltContentAddress = XLUtil.colRowNum2Str(srcCr.min_col, rowIndex)
            brContentAddress = XLUtil.colRowNum2Str(srcCr.max_col, rowIndex + eachSalaryContentRowCnt - 1)
            contentRange = "%s:%s" % (ltContentAddress, brContentAddress)
            util.copyValue(sheetTile, contentRange, dstSheetTitle, "A%s" % (copyRowCnt + 1),
                           dataOnly=True, autoSave=True)
            copyRowCnt += eachSalaryContentRowCnt

            copyRowCnt += eachGap
        dstTotalRangStr = 'A1:%s%s' % (XLUtil.colNum2Str(headWidth), copyRowCnt)
        dstSheet.print_area = dstTotalRangStr  # 打印区域
        dstSheet.page_setup.scale = 50
        # CommonUtil.printLog("打印区域是: %s" % dstSheet.print_area)
        util.save()
        return dstSheetTitle, dstTotalRangStr

    def sendEmailByImage(self, imageDirPath: str, emailConfigIniPath: str,
                         durationSec: int = -1,
                         needConfirm: bool = False) -> int:
        """
        根据工资截图信息, 批量发送邮件
        :param imageDirPath: 截图所在目录绝对路径, 要去目录中存在工资截图, 文件名格式: 姓名_邮箱.png , 如: 张三_123456@qq.com
        :param durationSec: 邮件发送间隔, 单位:秒, 正数有效, 若传其他值, 则会从 config.ini 中读取配置
        :param needConfirm: 发送前是否需要二次确认
        :param emailConfigIniPath: 发件人邮箱信息配置参数, 要求内部包含字段如下:
            [config]
            # 邮箱服务器, 默认使用qq邮箱
            smtpServer = smtp.qq.com

            # 服务器端口号
            smtpPort = 587

            # 发件人邮箱
            senderEmail = 4***06@qq.com

            # 邮箱密码或者授权码
            senderPwd = dflg***mdhcagc
        @return 发送成功的数量
        """
        successCount: int = 0  # 发送成功的数量
        CommonUtil.printLog(
            '批量发送工资截图到邮箱工具,连续发送间隔时长durationSec=%s, imageDirPath=%s' % (durationSec, imageDirPath))
        # 根据截图信息提取员工信息列表
        employeeList = list()  # 元素类型为: EmployeeSalaryInfo
        if not FileUtil.isDirFile(imageDirPath):
            NetUtil.push_to_robot(f'发送工资条邮件失败:您输入的不是目录路径, 请重新输入目录路径')
            return successCount

        imageList = FileUtil.listAllFilePath(imageDirPath)
        if len(FileUtil.listAllFilePath(imageDirPath)) == 0:
            NetUtil.push_to_robot(f'发送工资条邮件失败:该路径下并无文件, 请重新输入目录路径:{imageDirPath}')
            return successCount

            # 提取路径下的截图信息
        for imagePath in imageList:
            _, name, ext = FileUtil.getFileName(imagePath)
            if ext.lower() == 'png':
                employeeList.append(EmployeeSalaryInfo(imagePath))

        if len(employeeList) <= 0:
            CommonUtil.printLog('该路径下并无png截图, 请重新输入目录路径')
            NetUtil.push_to_robot(f'发送工资条邮件失败:该路径下并无png截图, 请重新输入目录路径:{imageDirPath}')
            return successCount

            # 发送成功的记录,每行一条记录, 格式:  xxxx 姓名 邮箱地址
        sendMailSuccessLog = '%s/send_mail_success_history.txt' % imageDirPath
        skipNameMailList = []  # 发送成功的邮件无需再次发送, 格式: 姓名_邮箱地址
        if FileUtil.isFileExist(sendMailSuccessLog):
            lines = FileUtil.readFile(sendMailSuccessLog)
            for line in lines:
                arr = line.strip().split(' ')
                if len(arr) < 3:
                    CommonUtil.printLog('sendMailSuccessLog invalid: %s' % line)
                    continue
                skipNameMailList.append('%s_%s' % (arr[1].strip(), arr[2].strip()))
        NetUtil.push_to_robot(f'已发送成功,无需再次发送的邮件共:{len(skipNameMailList)} 封')
        CommonUtil.printLog('具体列表如下:\n\t%s' % '\n\t'.join(skipNameMailList))

        continueRun: bool = not needConfirm or CommonUtil.get_input_info(
            f"以上是已发送成功过的邮件 {len(skipNameMailList)} 条(无需再次发送), 否继续执行, yes/no (默认:yes): ",
            "yes").lower() == "yes"
        if not continueRun:
            CommonUtil.printLog('退出程序')
            exit(1)

        CommonUtil.printLog('目录下的图片如下:')
        for item in employeeList:
            CommonUtil.printLog('\t%s' % item.imageName)

        imageSize = len(employeeList)
        isImageSizeValid: bool = not needConfirm or CommonUtil.get_input_info(
            f"\n共 {imageSize} 张图片, 是否正确, yes/no (默认:yes): ", "yes").lower() == "yes"
        if isImageSizeValid:
            sendEmailNow: bool = not needConfirm or CommonUtil.get_input_info("\n是否开始发送邮件, yes/no (默认:yes): ",
                                                                              "yes").lower() == "yes"
            # 显示属性修改提示
            # 默认使用当前目录下的 config.ini 文件路径
            # curDirPath = os.path.abspath(os.path.dirname(__file__))
            # configPath = '%s/config.ini' % curDirPath

            configParser = NewConfigParser().initPath(emailConfigIniPath)
            mailConfig = configParser.getSectionItems('mail')
            if sendEmailNow:
                CommonUtil.printLog('准备发送邮件...')
                subject = mailConfig['subject']

                if durationSec <= 0:
                    durationSec = int(mailConfig['delaySec'])

                mailUtil: MailUtil = None
                try:
                    for employeeInfo in employeeList:
                        # qq要求使用邮件授权码, 依次传入发件人邮箱, 密码, 发送服务器地址, 服务器端口
                        mailUtil = MailUtil(mailConfig)

                        path = employeeInfo.imagePath
                        name, email = employeeInfo.name, employeeInfo.email
                        curNameEmail = '%s_%s' % (name, email)
                        if curNameEmail in skipNameMailList:
                            CommonUtil.printLog('%s 已发送过邮件,无需重复发送,跳过' % curNameEmail)
                            successCount += 1
                            continue

                        # 正文中插入图片, 测试后,该文件不会再显示再附件中(qq邮箱)
                        # 1. 添加图片附件
                        mailUtil.obtainMailMsg(True)  # 每个人新建一个消息
                        cid = mailUtil.addAttachFile(path, "image")

                        # 2. 插入image标签
                        mailUtil.setMailMsg('<html><body>'
                                            + '<h2> %s,你好:</h2>' % name
                                            + '<p>这是您本月工资条,请查收</p>'
                                            + '<p><img src="cid:%s"></p>' % cid
                                            + '</body></html>',
                                            subject=f"{subject}-{name}")

                        # 添加普通文件附件
                        # mailUtil.addAttachFile("/Users/***/Desktop/test.py", "file")
                        # 发送到指定邮箱
                        CommonUtil.printLog("正在发送邮件给 %s %s" % (name, email))
                        if name is not None and len(name) > 0 and email is not None and '@' in email:
                            sendSuccess = mailUtil.sendTo([email])
                            if sendSuccess:
                                successCount += 1
                                FileUtil.append2File(sendMailSuccessLog, "成功发送邮件给 %s %s" % (name, email))
                            else:
                                NetUtil.push_to_robot(f'发送 {name} 工资条邮件失败,图片:{employeeInfo.imagePath}')

                        mailUtil.quit()
                        mailUtil = None
                        if durationSec > 0:
                            time.sleep(durationSec)

                except Exception as e:
                    NetUtil.push_to_robot(f'邮件发送出现异常, 退出发送, 请排查: {e}')
                finally:
                    if mailUtil is not None:
                        mailUtil.quit()
            NetUtil.push_to_robot(f'发送邮件完成,成功: {successCount}/{imageSize}')
            return successCount

    @log_time_consume()
    def take_shot(self, sheetTile: str, headerRangeStr: str, total_range: str,
                  eachSalaryContentRowCnt: int = 1,
                  emailColLetter: str = None, nameColLetter: str = None) -> dict:
        """
        拆分工资汇总表为每个人独立一个工资条(固定的标题栏+个人工资信息+gap空白分隔行)
        :param sheetTile: 待截图的工资表名, 如: "sheet1"
        :param headerRangeStr:工资条固定标题栏区域,如: "A1:Z2"
        :param total_range: 总区域(含标题和内容), 如 "A1:Z50"
        :param eachSalaryContentRowCnt: 每个人的工资信息行数,默认1行
        :param emailColLetter: 邮箱信息所在列号
        :param nameColLetter: 名字信息所在列号
        :return:  map(截图名称, 截图路径)  截图名称格式: {姓名}_{邮箱}
        """
        img_dict = {}  # 工资条信息 key-姓名+邮箱地址 value-图片路径
        continueMode: bool = "true" == self.configParser.getSectionItems('config')["continueMode"].lower()
        if not FileUtil.isFileExist(self.imgDirPath):
            FileUtil.createFile(self.imgDirPath, not continueMode)  # continueMode下不重建目录
            if not FileUtil.isFileExist(self.imgDirPath):
                CommonUtil.printLog(f'截图缓存目录创建失败, 请手动创建后重试 {self.imgDirPath}')

        headWidth, headHeight = XLUtil.getRangeWH(headerRangeStr)  # 标题行
        srcCr = cell_range.CellRange(range_string=total_range)

        min_row = srcCr.min_row
        max_row = srcCr.max_row
        min_col = srcCr.min_col
        max_col = srcCr.max_col

        _xl_shot_util = ExcelTakeShotUtil(self.xl_path, sheetTile, img_cache_dir=self.imgDirPath)
        try:
            for rowIndex in range(min_row, max_row + 1, eachSalaryContentRowCnt + headHeight):
                # 复制个人工资条信息
                ltContentAddress = XLUtil.colRowNum2Str(min_col, rowIndex)
                brContentAddress = XLUtil.colRowNum2Str(max_col, rowIndex + eachSalaryContentRowCnt + headHeight - 1)
                area = "%s:%s" % (ltContentAddress, brContentAddress)  # 个人工资条区域,包含标题栏和内容

                # 获取姓名和邮箱地址
                name = _xl_shot_util.getCellValue("%s%s" % (nameColLetter, rowIndex + headHeight))
                email = _xl_shot_util.getCellValue("%s%s" % (emailColLetter, rowIndex + headHeight))
                if CommonUtil.isNoneOrBlank(name):
                    CommonUtil.printLog(f" {nameColLetter}{rowIndex} 姓名为空,取消其工资条截图,请注意复查")
                    self.fail_pic_cnt += 1
                elif CommonUtil.isNoneOrBlank(email):
                    CommonUtil.printLog(f' {emailColLetter}{rowIndex} 员工 {name} 并未提供邮箱地址, 取消其工资条截图')
                    self.fail_pic_cnt += 1
                else:
                    name = name.strip()
                    email = email.strip()
                    # 截图
                    imgName = f"{name}_{email}"
                    imgPath = FileUtil.recookPath(f"{self.imgDirPath}/{imgName}.png")
                    success, path = _xl_shot_util.take_shot(area, imgPath, continueMode)
                    if not success:
                        CommonUtil.printLog(f'{name} 的工资条截图失败, 请手动截图')
                        self.fail_pic_cnt += 1
                    else:
                        # CommonUtil.printLog("截图成功,图片地址:%s" % path)
                        self.success_pic_cnt += 1
                        img_dict[imgName] = path
            CommonUtil.printLog(f'截图结果共{len(img_dict)}张, 失败{self.fail_pic_cnt}张, 存储路径为:{self.imgDirPath}')
        except Exception as e:
            CommonUtil.printLog(f'截图出现异常, 请排查: {e}')
        finally:
            _xl_shot_util.close()
        return img_dict
